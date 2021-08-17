import openpyxl as xl
from user import User
"""message = "This is a simple message"  # string
day = 9  # integer
price = 19.99  # float
valid = True  # or can be False Boolean
list_of_months = ['January', 'February', 'March', 'April']  # list
set_of_day = {10, 15, 35, 90}  # set
days_and_units = {'days': 20, "unit": "hours"}  # dictionary"""


"""inv_file = xl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]
product_per_supplier = {}
total_value_per_supplier = {}
product_under_ten = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    # new col
    inventory_price = product_list.cell(product_row, 5)

    # product per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] = product_per_supplier.get(supplier_name) + 1
    else:
        product_per_supplier[supplier_name] = 1

    # calculate total value per supplier
    if supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calculate The product under 10 inventory
    if inventory < 10:
        product_under_ten[product_number] = inventory

    # add value for each row :)
    inventory_price.value = inventory * price


print(product_per_supplier)
print(total_value_per_supplier)
print(product_under_ten)
# save the change xlsx file in new one
inv_file.save("inventory-new.xlsx")"""

muhmdreza = User("mgh2711@gmail.com", "muhmdrezA", 12345, "developer")
muhmdreza.user_get_info()